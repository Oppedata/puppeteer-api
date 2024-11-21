from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import openpyxl

# ตั้งค่า ChromeDriver
service = Service('C:\\chromedriver\\chromedriver.exe')  # แก้ไข path ตามที่เก็บ ChromeDriver
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# เปิดหน้าเว็บ
url = "http://182.52.47.34/#/landing"
driver.get(url)

try:
    # ดึงข้อมูลจาก Power
    power_pea = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//div[@class="val val1"]'))
    ).text
    power_mahidol = driver.find_element(By.XPATH, '//div[@class="val val2"]').text
    power_solar = driver.find_element(By.XPATH, '//div[@class="val val3"]').text

    # ดึงข้อมูลจาก CO2 Saving
    co2_from_cod = driver.find_element(By.XPATH, '//div[@class="val1"]/p[@class="valtxt"]').text
    co2_this_month = driver.find_element(By.XPATH, '//div[@class="val2"]/p[@class="valtxt"]').text

    # ดึงข้อมูลจาก Energy
    energy_this_month = driver.find_element(By.XPATH, '//div[@class="val1"]/p[@class="valtxt"]').text
    energy_this_year = driver.find_element(By.XPATH, '//div[@class="val2"]/p[@class="valtxt"]').text
    energy_from_cod = driver.find_element(By.XPATH, '//div[@class="val3"]/p[@class="valtxt"]').text

    # สร้าง DataFrame สำหรับแต่ละหมวด
    df_power = pd.DataFrame({
        "Category": ["PEA", "Mahidol", "Solar"],
        "Value (kW)": [power_pea, power_mahidol, power_solar]
    })

    df_co2 = pd.DataFrame({
        "Category": ["CO2 Saving From COD", "CO2 Saving This Month"],
        "Value (Ton CO2e)": [co2_from_cod, co2_this_month]
    })

    df_energy = pd.DataFrame({
        "Category": ["Energy This Month", "Energy This Year", "Energy From COD"],
        "Value (kWh)": [energy_this_month, energy_this_year, energy_from_cod]
    })

    # บันทึกข้อมูลลงใน Excel โดยเพิ่มข้อมูลในแถวใหม่
    output_file = "output_data.xlsx"

    # ตรวจสอบว่าไฟล์มีอยู่แล้วหรือไม่
    try:
        # โหลดไฟล์ Excel หากมีอยู่
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            startrow_power = writer.sheets["Power (kW)"].max_row
            startrow_co2 = writer.sheets["CO2 Saving"].max_row
            startrow_energy = writer.sheets["Energy (kWh)"].max_row

            df_power.to_excel(writer, sheet_name="Power (kW)", index=False, header=False, startrow=startrow_power)
            df_co2.to_excel(writer, sheet_name="CO2 Saving", index=False, header=False, startrow=startrow_co2)
            df_energy.to_excel(writer, sheet_name="Energy (kWh)", index=False, header=False, startrow=startrow_energy)

    except FileNotFoundError:
        # หากไฟล์ไม่พบ ให้สร้างไฟล์ใหม่พร้อม Header
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_power.to_excel(writer, sheet_name="Power (kW)", index=False)
            df_co2.to_excel(writer, sheet_name="CO2 Saving", index=False)
            df_energy.to_excel(writer, sheet_name="Energy (kWh)", index=False)

    print(f"ข้อมูลถูกเพิ่มลงในไฟล์ '{output_file}' เรียบร้อยแล้ว!")

except Exception as e:
    print(f"เกิดข้อผิดพลาด: {e}")

# ปิดเบราว์เซอร์
driver.quit()
